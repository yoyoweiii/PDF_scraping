import pdfplumber
import openpyxl
import os
import json
from pdfminer.high_level import extract_text

def extract_tables_and_text(row_n, pdf_path):
    print(pdf_path," is processing.")
    content = ""
    table_content= []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                #print(f"Page {page_num + 1}:")
                content+=f"Page {page_num + 1}:\n"
                text = page.extract_text()
                #print("Text Content:")
                content += "\nText Content:\n"
                #print(text)
                content += text
                for table_num, table in enumerate(page.extract_tables()):
                    print(f"Table {table_num + 1}:")
                    for r in table:
                        table_content.append(r)
            json_table = json.dumps(table_content)
        sheet.cell(row=row_n, column=14, value=json_table)
        sheet.cell(row=row_n, column=10, value=content)
        print(pdf_path, " is done.")
    except:
        print("not valid path")
#pdf_path = r"C:\Users\TSECDC01\Desktop\SDxResearch\Line Class\CA\IEM OA4.pdf"
excel_file = 'file_plant_folder-match.xlsx'
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active
#pdf_path = r"C:\Users\TSECDC01\Desktop\SDxResearch\Line Class\CA\EDC AAG.pdf"
row_n = 2
for row in sheet.iter_rows():
    pdf_path = os.path.join(row[3].value, row[1].value)
    extract_tables_and_text( row_n, pdf_path)
    row_n+=1
workbook.save(excel_file)