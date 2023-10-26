import pdfplumber
import openpyxl
from openpyxl import Workbook
import os
import json
from pdfminer.high_level import extract_text

workbook = Workbook()
sheet = workbook.active
pdf_path = r"C:\Users\TSECDC01\Desktop\SDxResearch\Line Class\PP2\PP3 Pipe Spec.pdf"
excel_file = 'PP3 Pipe Spec(116 sheets).xlsx'
pdf = pdfplumber.open(pdf_path)
num_pages = len(pdf.pages)

for page_num, page in enumerate(pdf.pages):
    row_num=1
    sheet_name = f"Sheet{page_num+1}"
    sheet = workbook.create_sheet(title=sheet_name)
    print(f"Page {page_num}:")
    text = page.extract_text()
    lines = text.split('\n')
    for row in lines:
        sheet.cell(row=row_num, column=1, value=row)
        row_num = row_num + 1
    #workbook.save(value=text)
    for table_num, table in enumerate(page.extract_tables()):
        #print(f"Table {table_num + 1}:")
        for r in table:
            #print(r)
            json_table = json.dumps(r)
            sheet.cell(row=row_num, column=1, value=json_table)
            row_num = row_num + 1

print(pdf_path, " is done.")
# Save the workbook
workbook.save(excel_file)
workbook.close()