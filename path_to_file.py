import os
import openpyxl
def list_all_files_in_directory(directory):
    all_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith("pdf"):
                file_path = os.path.join(root, file)
                sheet.cell(row=sheet.max_row+1, column=1, value=sheet.max_row)
                sheet.cell(row=sheet.max_row, column=2, value=file)
                x = file_path.split("\\")
                print(x[6])
                sheet.cell(row=sheet.max_row, column=3, value=x[6])
                sheet.cell(row=sheet.max_row, column=4, value=root)
                all_files.append(file_path)

    return all_files

if __name__ == "__main__":
    current_directory = os.getcwd()
    excel_file_path="file_plant_folder-match.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    all_files = list_all_files_in_directory(current_directory)
    n=1
    for file_path in all_files:
        #print(file_path)
        n+=1
    print(n)
    workbook.save(excel_file_path)